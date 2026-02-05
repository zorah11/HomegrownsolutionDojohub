# COMPLETE DATA MOVEMENT - BOTH SYSTEMS WITH ACTUAL VALUES
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

# Note: This assumes all the required variables are already in the notebook's namespace
# This needs to be run within the Jupyter notebook environment

print("="*80)
print("CREATING COMPLETE SYSTEM DATA MOVEMENT EXCEL")
print("="*80)

wb = openpyxl.Workbook()

# ============================================================================
# SHEET 1: YIELD PREDICTION DATA FLOW
# ============================================================================
ws1 = wb.active
ws1.title = "Yield_Prediction_Flow"

headers1 = ['Step', 'Transformation', 'Date', 'Season', 'NDVI_Mean', 'NDVI_Std', 'Images', 'Temp_C', 'Rain_mm', 'Feature_Sample', 'Output']
for col, h in enumerate(headers1, 1):
    cell = ws1.cell(1, col, h)
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

row = 2

# Step 1: Raw NDVI (show all 46 observations)
print(f"[1/7] Extracting {len(df_timeseries)} raw NDVI observations...")
for idx, r in df_timeseries.iterrows():
    ws1.cell(row, 1, "1-Raw_NDVI")
    ws1.cell(row, 2, "Satellite → NDVI timeseries")
    ws1.cell(row, 3, str(r['date']))
    ws1.cell(row, 4, r['cycle_name'])
    ws1.cell(row, 5, round(r['ndvi_mean'], 3))
    ws1.cell(row, 6, round(r['ndvi_std'], 3))
    ws1.cell(row, 7, int(r['image_count']))
    ws1.cell(row, 11, f"{len(df_timeseries)} observations")
    row += 1

# Step 2: Raw Weather (show all 46)
print(f"[2/7] Extracting {len(df_weather)} weather records...")
for idx, r in df_weather.iterrows():
    ws1.cell(row, 1, "2-Weather")
    ws1.cell(row, 2, "ERA5/CHIRPS → Weather data")
    ws1.cell(row, 3, str(r['date']))
    ws1.cell(row, 4, r['cycle_name'])
    ws1.cell(row, 8, round(r['temperature'], 1))
    ws1.cell(row, 9, round(r['precipitation'], 1))
    ws1.cell(row, 11, f"{len(df_weather)} observations")
    row += 1

# Step 3: Combined Data (46)
print(f"[3/7] Merging to {len(df_combined)} combined records...")
for idx, r in df_combined.iterrows():
    ws1.cell(row, 1, "3-Combined")
    ws1.cell(row, 2, "NDVI + Weather merge")
    ws1.cell(row, 3, str(r['date']))
    ws1.cell(row, 4, r['cycle_name'])
    ws1.cell(row, 5, round(r['ndvi_mean'], 3))
    ws1.cell(row, 8, round(r['temperature'], 1))
    ws1.cell(row, 9, round(r['precipitation'], 1))
    ws1.cell(row, 11, f"{len(df_combined)} combined")
    row += 1

# Step 4: Season Features (5 seasons × 29 features = 145 rows)
print(f"[4/7] Generating {len(df_features_enhanced)} season-level features...")
for idx, r in df_features_enhanced.iterrows():
    sample_feats = [c for c in r.index if 'peak' in c or 'mean' in c][:3]
    feat_cols = [f"{c}={r[c]:.2f}" for c in sample_feats if pd.notna(r[c])]
    
    ws1.cell(row, 1, "4-Season_Features")
    ws1.cell(row, 2, "Seasonal aggregation (29 features)")
    ws1.cell(row, 4, r['cycle_name'])
    ws1.cell(row, 10, ", ".join(feat_cols))
    ws1.cell(row, 11, f"145 features total")
    row += 1

# Step 5: Feature Selection (top 5 selected)
print(f"[5/7] Selecting top {len(top_feature_names)} features...")
top_features = feature_importance.head(5)
for idx, (feat, importance) in enumerate(zip(top_features['feature'], top_features['importance'])):
    ws1.cell(row, 1, "5-Feature_Selection")
    ws1.cell(row, 2, "RandomForest Feature Importance")
    ws1.cell(row, 10, f"{feat} (importance={importance:.3f})")
    ws1.cell(row, 11, f"{len(top_feature_names)} features selected")
    row += 1

# Step 6: Cross-Validation Results (5 folds)
print(f"[6/7] Running {len(y_train)} cross-validation folds...")
for i, (actual, predicted) in enumerate(zip(y_train, y_pred_cv)):
    error = abs(actual - predicted)
    error_pct = (error / actual) * 100
    ws1.cell(row, 1, "6-Cross_Validation")
    ws1.cell(row, 2, "Leave-One-Out CV")
    ws1.cell(row, 4, f"Season {i+1}")
    ws1.cell(row, 10, f"Actual={actual:.0f}, Pred={predicted:.0f}, Error={error:.0f} ({error_pct:.1f}%)")
    ws1.cell(row, 11, f"MAE={mae_cv:.1f} kg, R²={r2_cv:.3f}")
    row += 1

# Step 7: Final Prediction
print(f"[7/7] Making final prediction...")
pred_std = (confidence_interval[1] - confidence_interval[0]) / (2 * 1.96)
ws1.cell(row, 1, "7-Final_Prediction")
ws1.cell(row, 2, "Model → Current Season")
ws1.cell(row, 4, current_cycle_name)
ws1.cell(row, 10, f"Yield = {predicted_yield:.1f} kg")
ws1.cell(row, 11, f"Confidence: ±{pred_std:.1f} kg ({confidence_interval[0]:.0f}-{confidence_interval[1]:.0f})")

print(f"   ✓ Sheet 1: {row} rows showing complete yield prediction flow\n")

# ============================================================================
# SHEET 2: FARMER MONITORING FLOW
# ============================================================================
ws2 = wb.create_sheet("Farmer_Monitoring_Flow")

headers2 = ['Farmer', 'Crop', 'Acres', 'GPS_Lat', 'GPS_Lon', 'Soil_pH', 'Nitrogen_g_kg', 'SOC_g_kg', 'Moisture_Status', 'Anomaly_Status']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(1, col, h)
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")

row = 2

print("[PART 2] Farmer Monitoring System - ALL FARMERS WITH ACTUAL VALUES")
print("-" * 80)

total_farmers = 0
total_crops = 0
total_acreage = 0

for farmer_name, farms in farmer_portfolio.items():
    total_farmers += 1
    farmer_total_acres = sum(f['size'] for f in farms)
    total_crops += len(farms)
    total_acreage += farmer_total_acres
    
    print(f"\nFarmer {total_farmers}: {farmer_name} - {len(farms)} crops, {farmer_total_acres} acres")
    print(f"  Querying satellite data...")
    
    # Get actual soil data directly from SoilGrids ONCE per farmer
    try:
        farm_poly = get_farm_geometry(farms[0]['coords'])
        point = farm_poly.centroid()
        
        ph_img = ee.Image("projects/soilgrids-isric/phh2o_mean").select('phh2o_0-5cm_mean')
        nitrogen_img = ee.Image("projects/soilgrids-isric/nitrogen_mean").select('nitrogen_0-5cm_mean')
        soc_img = ee.Image("projects/soilgrids-isric/soc_mean").select('soc_0-5cm_mean')
        
        combined = ee.Image.cat([ph_img, nitrogen_img, soc_img])
        stats = combined.reduceRegion(reducer=ee.Reducer.mean(), geometry=point, scale=250).getInfo()
        
        if stats.get('phh2o_0-5cm_mean') is not None:
            soil_ph = stats.get('phh2o_0-5cm_mean') / 10
            soil_n = stats.get('nitrogen_0-5cm_mean') / 100
            soil_soc = stats.get('soc_0-5cm_mean') / 10
            print(f"  ✓ Soil: pH={soil_ph:.2f}, N={soil_n:.2f}, SOC={soil_soc:.2f}")
        else:
            soil_ph = soil_n = soil_soc = 0
            print(f"  ✗ Soil data unavailable for location")
    except Exception as e:
        soil_ph = soil_n = soil_soc = 0
        print(f"  ✗ Soil error: {str(e)[:50]}")
    
    try:
        moisture_data = get_soil_moisture_trend(farmer_name)
        moisture_status = "ALERT" if "ALERT" in moisture_data else "Adequate"
        print(f"  ✓ Moisture: {moisture_status}")
    except Exception as e:
        moisture_status = "Unknown"
        print(f"  ✗ Moisture error")
    
    try:
        anomaly_data = detect_pest_anomaly(farmer_name)
        anomaly_status = "WARNING" if ("WARNING" in anomaly_data or "ALERT" in anomaly_data) else "Healthy"
        print(f"  ✓ Anomaly: {anomaly_status}")
    except Exception as e:
        anomaly_status = "Unknown"
        print(f"  ✗ Anomaly error")
    
    # Each crop for this farmer - ONE ROW PER CROP
    for farm in farms:
        crop_name = farm['crop']
        crop_size = farm['size']
        coords = farm['coords']
        
        ws2.cell(row, 1, farmer_name)
        ws2.cell(row, 2, crop_name)
        ws2.cell(row, 3, crop_size)
        ws2.cell(row, 4, round(coords[0][0], 4))
        ws2.cell(row, 5, round(coords[0][1], 4))
        ws2.cell(row, 6, round(soil_ph, 2) if soil_ph > 0 else "N/A")
        ws2.cell(row, 7, round(soil_n, 2) if soil_n > 0 else "N/A")
        ws2.cell(row, 8, round(soil_soc, 1) if soil_soc > 0 else "N/A")
        ws2.cell(row, 9, moisture_status)
        ws2.cell(row, 10, anomaly_status)
        row += 1

print(f"\n   ✓ Sheet 2: {row-1} rows with ACTUAL NUMERIC VALUES\n")

# Save file
filename = "Complete_System_Data_Movement.xlsx"
wb.save(filename)

print("="*80)
print(f"✅ EXPORT COMPLETE: {filename}")
print("="*80)
print(f"📊 Sheet 1: Yield Prediction - Complete data transformation flow")
print(f"   • {len(df_timeseries)} raw NDVI → {len(df_weather)} weather → {len(df_combined)} combined")
print(f"   • {len(df_features_enhanced)} seasons × 29 features → {len(top_feature_names)} selected")
print(f"   • CV: MAE={mae_cv:.1f} kg, R²={r2_cv:.3f}")
print(f"   • Final: {predicted_yield:.1f} ± {pred_std:.1f} kg")
print(f"\n📊 Sheet 2: Farmer Monitoring - Complete monitoring flow")
print(f"   • {total_farmers} farmers: {total_crops} crops, {total_acreage} acres total")
print(f"   • pH, N, SOC columns showing ACTUAL NUMERIC VALUES")
print("="*80)
